import json
from pathlib import Path
from collections import Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def load_all_quotes(output_dir: Path) -> list[dict]:
    """Load all quotes from *_quotes.json files in output directory."""
    all_quotes = []
    quote_files = sorted(output_dir.glob("*_quotes.json"))
    
    for file_path in quote_files:
        with open(file_path, "r", encoding="utf-8") as f:
            quotes = json.load(f)
            all_quotes.extend(quotes)
    
    return all_quotes


def transform_quotes_to_dataframe(quotes: list[dict]) -> pd.DataFrame:
    """Transform quotes list to DataFrame with proper column order."""
    rows = []
    
    for idx, quote in enumerate(quotes, start=1):
        row = {
            "id": idx,
            "speaker": quote.get("speaker", ""),
            "speaker_function": quote.get("speaker_function", ""),
            "speaker_expertise": ", ".join(quote.get("speaker_expertise", [])),
            "text": quote.get("text", ""),
            "text_ko": quote.get("text_ko", ""),
            "text_zh": quote.get("text_zh", ""),
            "text_es": quote.get("text_es", ""),
            "timestamp": quote.get("timestamp", ""),
            "context": quote.get("context", ""),
            "vocabulary_highlights": ", ".join(quote.get("vocabulary_highlights", [])),
            "topics": ", ".join(quote.get("topics", [])),
            "difficulty_level": quote.get("difficulty_level", "")
        }
        rows.append(row)
    
    # Define column order
    columns = [
        "id", "speaker", "speaker_function", "speaker_expertise",
        "text", "text_ko", "text_zh", "text_es",
        "timestamp", "context", "vocabulary_highlights", "topics", "difficulty_level"
    ]
    
    df = pd.DataFrame(rows, columns=columns)
    return df


def create_formatted_excel(df: pd.DataFrame, output_path: Path):
    """Create formatted Excel file with auto-width, filters, frozen header, and bold header."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"
    
    # Write DataFrame to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Bold header row
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # Wrap text for long content columns
                if c_idx in [5, 6, 7, 8, 10]:  # text columns and context
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-width columns (with max width limit for readability)
    max_widths = {
        "id": 6,
        "speaker": 25,
        "speaker_function": 18,
        "speaker_expertise": 35,
        "text": 60,
        "text_ko": 50,
        "text_zh": 50,
        "text_es": 50,
        "timestamp": 12,
        "context": 60,
        "vocabulary_highlights": 35,
        "topics": 30,
        "difficulty_level": 15
    }
    
    for col_idx, column in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        
        # Calculate width based on content (sample first 100 rows)
        max_length = len(str(column))  # Start with header length
        for row_idx in range(2, min(len(df) + 2, 102)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, min(len(str(cell_value)), 100))
        
        # Apply width with limits
        width = min(max_length + 2, max_widths.get(column, 50))
        ws.column_dimensions[col_letter].width = width
    
    # Add filters to header row
    ws.auto_filter.ref = ws.dimensions
    
    # Freeze first row (header)
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(output_path)


def print_summary(df: pd.DataFrame, quotes: list[dict]):
    """Print detailed summary statistics."""
    print("\n" + "=" * 60)
    print("EXPORT SUMMARY")
    print("=" * 60)
    
    # Total quotes
    print(f"\nTotal quotes: {len(df)}")
    
    # Total speakers
    unique_speakers = df["speaker"].nunique()
    print(f"Total speakers: {unique_speakers}")
    
    # Speaker functions with counts
    print(f"\nSpeaker functions:")
    function_counts = df["speaker_function"].value_counts()
    for func, count in function_counts.items():
        print(f"  - {func}: {count}")
    
    # Topics with counts (need to split and count individually)
    print(f"\nTopics:")
    all_topics = []
    for quote in quotes:
        all_topics.extend(quote.get("topics", []))
    topic_counts = Counter(all_topics)
    for topic, count in topic_counts.most_common():
        print(f"  - {topic}: {count}")
    
    # Difficulty levels with counts
    print(f"\nDifficulty levels:")
    difficulty_counts = df["difficulty_level"].value_counts()
    for level, count in difficulty_counts.items():
        print(f"  - {level}: {count}")
    
    # Languages
    print(f"\nLanguages: English, Korean, Chinese, Spanish")
    
    # Check translation coverage
    ko_count = df["text_ko"].notna().sum() - (df["text_ko"] == "").sum()
    zh_count = df["text_zh"].notna().sum() - (df["text_zh"] == "").sum()
    es_count = df["text_es"].notna().sum() - (df["text_es"] == "").sum()
    print(f"  - English (text): {len(df)} quotes")
    print(f"  - Korean (text_ko): {ko_count} quotes")
    print(f"  - Chinese (text_zh): {zh_count} quotes")
    print(f"  - Spanish (text_es): {es_count} quotes")
    
    print("\n" + "=" * 60)


def main():
    base_dir = Path(__file__).parent
    output_dir = base_dir / "output"
    
    # Check if output directory exists
    if not output_dir.exists():
        print("Error: output/ directory not found.")
        return
    
    # Load all quotes
    print("Loading quote files...")
    quotes = load_all_quotes(output_dir)
    
    if not quotes:
        print("No quotes found in *_quotes.json files.")
        return
    
    print(f"Loaded {len(quotes)} quotes from output/ folder.")
    
    # Transform to DataFrame
    print("Transforming data...")
    df = transform_quotes_to_dataframe(quotes)
    
    # Export to Excel
    excel_path = output_dir / "quotes_complete.xlsx"
    print(f"Creating Excel file: {excel_path}")
    create_formatted_excel(df, excel_path)
    print(f"  -> Saved: {excel_path}")
    
    # Export to CSV
    csv_path = output_dir / "quotes_complete.csv"
    print(f"Creating CSV file: {csv_path}")
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"  -> Saved: {csv_path}")
    
    # Print summary
    print_summary(df, quotes)


if __name__ == "__main__":
    main()
