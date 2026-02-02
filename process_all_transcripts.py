"""
Full pipeline for processing Lenny's Podcast transcripts.
Extracts quotes, translates to multiple languages, and exports to Excel.

Features:
- Batch processing (50 files at a time)
- Resume capability with checkpoints
- Progress bar and ETA
- Cost tracking
- Error handling with retries
- Detailed logging
"""

import os
import json
import time
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional
from dotenv import load_dotenv
from openai import OpenAI
from anthropic import Anthropic

# Load environment variables
load_dotenv('.env.local')

# ============================================================================
# CONFIGURATION
# ============================================================================

BATCH_SIZE = 50
BATCH_WAIT_MINUTES = 10  # Wait between batches for rate limit protection
ERROR_RETRY_WAIT_MINUTES = 5
MAX_RETRIES = 3
API_DELAY_SECONDS = 2

# Cost estimates (per file, approximate)
# Extract: ~4K tokens input, ~1K output per transcript
# Translate: ~3 API calls per quote, ~10 quotes per file average
COST_PER_FILE_EXTRACT_USD = 0.02  # Claude extraction
COST_PER_FILE_TRANSLATE_USD = 0.03  # OpenAI + Claude translation
COST_PER_FILE_TOTAL_USD = COST_PER_FILE_EXTRACT_USD + COST_PER_FILE_TRANSLATE_USD

# Time estimates (per file, including delays)
SECONDS_PER_FILE_EXTRACT = 15  # API call + delay
SECONDS_PER_QUOTE_TRANSLATE = 8  # 3 API calls with delays
AVG_QUOTES_PER_FILE = 10

# Model configurations
OPENAI_MODEL = "gpt-5-mini"
ANTHROPIC_MODEL = "claude-sonnet-4-20250514"

# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging(base_dir: Path) -> logging.Logger:
    """Setup logging to both console and file."""
    logger = logging.getLogger("pipeline")
    logger.setLevel(logging.INFO)
    
    # Clear existing handlers
    logger.handlers = []
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_format = logging.Formatter('%(message)s')
    console_handler.setFormatter(console_format)
    logger.addHandler(console_handler)
    
    # File handler for errors
    error_log_path = base_dir / "errors.log"
    file_handler = logging.FileHandler(error_log_path, encoding="utf-8")
    file_handler.setLevel(logging.ERROR)
    file_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_format)
    logger.addHandler(file_handler)
    
    return logger


# ============================================================================
# CHECKPOINT MANAGEMENT
# ============================================================================

class CheckpointManager:
    """Manages pipeline checkpoints for resume capability."""
    
    def __init__(self, base_dir: Path):
        self.checkpoint_path = base_dir / "pipeline_checkpoint.json"
        self.data = self._load()
    
    def _load(self) -> dict:
        """Load checkpoint from file."""
        if self.checkpoint_path.exists():
            with open(self.checkpoint_path, "r", encoding="utf-8") as f:
                return json.load(f)
        return {
            "extracted_files": [],
            "translated_files": [],
            "errors": [],
            "start_time": None,
            "total_cost_usd": 0.0,
            "total_quotes_extracted": 0,
            "total_quotes_translated": 0
        }
    
    def save(self):
        """Save checkpoint to file."""
        with open(self.checkpoint_path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)
    
    def mark_extracted(self, filename: str, quote_count: int, cost: float):
        """Mark a file as extracted."""
        if filename not in self.data["extracted_files"]:
            self.data["extracted_files"].append(filename)
            self.data["total_quotes_extracted"] += quote_count
            self.data["total_cost_usd"] += cost
        self.save()
    
    def mark_translated(self, filename: str, quote_count: int, cost: float):
        """Mark a file as translated."""
        if filename not in self.data["translated_files"]:
            self.data["translated_files"].append(filename)
            self.data["total_quotes_translated"] += quote_count
            self.data["total_cost_usd"] += cost
        self.save()
    
    def add_error(self, error_msg: str):
        """Add an error to the checkpoint."""
        timestamp = datetime.now().isoformat()
        self.data["errors"].append(f"[{timestamp}] {error_msg}")
        self.save()
    
    def is_extracted(self, filename: str) -> bool:
        """Check if a file has been extracted."""
        return filename in self.data["extracted_files"]
    
    def is_translated(self, filename: str) -> bool:
        """Check if a file has been translated."""
        return filename in self.data["translated_files"]
    
    def set_start_time(self):
        """Set the pipeline start time."""
        if self.data["start_time"] is None:
            self.data["start_time"] = datetime.now().isoformat()
            self.save()
    
    def get_elapsed_time(self) -> timedelta:
        """Get elapsed time since start."""
        if self.data["start_time"]:
            start = datetime.fromisoformat(self.data["start_time"])
            return datetime.now() - start
        return timedelta(0)
    
    def clear(self):
        """Clear all checkpoint data."""
        self.data = {
            "extracted_files": [],
            "translated_files": [],
            "errors": [],
            "start_time": None,
            "total_cost_usd": 0.0,
            "total_quotes_extracted": 0,
            "total_quotes_translated": 0
        }
        self.save()


# ============================================================================
# PROGRESS BAR
# ============================================================================

def print_progress_bar(current: int, total: int, prefix: str = "", suffix: str = "", length: int = 40):
    """Print a progress bar to console."""
    percent = (current / total) * 100 if total > 0 else 0
    filled = int(length * current // total) if total > 0 else 0
    bar = "█" * filled + "░" * (length - filled)
    print(f"\r{prefix} |{bar}| {current}/{total} ({percent:.1f}%) {suffix}", end="", flush=True)


def format_time(seconds: float) -> str:
    """Format seconds into human-readable string."""
    if seconds < 60:
        return f"{int(seconds)}s"
    elif seconds < 3600:
        mins = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{mins}m {secs}s"
    else:
        hours = int(seconds // 3600)
        mins = int((seconds % 3600) // 60)
        return f"{hours}h {mins}m"


# ============================================================================
# API FUNCTIONS (from extract_quotes.py)
# ============================================================================

def load_file(file_path: Path) -> str:
    """Load and return contents of a text file."""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def extract_quotes_api(client: Anthropic, transcript: str, prompt: str) -> list:
    """Call Anthropic API to extract quotes from transcript."""
    message = client.messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=4096,
        messages=[
            {
                "role": "user",
                "content": f"{prompt}\n\nHere is the podcast transcript to analyze:\n\n{transcript}"
            }
        ]
    )
    
    response_text = message.content[0].text
    quotes = json.loads(response_text)
    return quotes


def enrich_quotes_with_speaker_info(quotes: list, speaker_profiles: dict) -> list:
    """Add speaker_function and speaker_expertise to each quote."""
    for quote in quotes:
        speaker_name = quote.get("speaker", "")
        
        if speaker_name in speaker_profiles:
            profile = speaker_profiles[speaker_name]
            quote["speaker_function"] = profile.get("function", "Leadership")
            quote["speaker_expertise"] = profile.get("expertise", [])
        else:
            quote["speaker_function"] = "Leadership"
            quote["speaker_expertise"] = []
    
    return quotes


# ============================================================================
# API FUNCTIONS (from translate_quotes_openai.py)
# ============================================================================

KOREAN_PROMPT = """Translate this English business quote into natural, conversational Korean that Korean professionals actually use. Use -요/-해요 ending. For technical terms/jargon, add English in parentheses like '호기심 루프(curiosity loop)'. Avoid stiff literal translation. Be natural and conversational. Return ONLY the Korean translation.

Quote to translate:
{text}"""

CLAUDE_PROMPTS = {
    "zh": """Translate this English business quote into natural, conversational Simplified Chinese that Chinese business professionals actually use.

CRITICAL RULES:
- Use natural spoken Chinese (口语化), not stiff/formal written Chinese
- Avoid literal word-for-word translation (避免翻译腔)
- For English jargon or technical terms, keep the English in parentheses after Chinese. Example: "好奇心循环(curiosity loop)"
- Use natural Chinese expressions and idioms
- Think: How would a Chinese person naturally say this in a business conversation?

Quote to translate:
{text}

Return ONLY the Chinese translation, nothing else.""",

    "es": """Translate this English business quote into natural, conversational Spanish that Spanish-speaking business professionals actually use.

CRITICAL RULES:
- Use natural spoken Spanish, not stiff/formal written Spanish
- Use a professional but approachable tone (tú/usted as appropriate for business)
- Avoid literal word-for-word translation
- For English jargon or technical terms, keep the English in parentheses after Spanish. Example: "bucle de curiosidad (curiosity loop)"
- Use natural Spanish expressions and idioms
- Think: How would a Spanish speaker naturally say this in a business conversation?

Quote to translate:
{text}

Return ONLY the Spanish translation, nothing else."""
}


def translate_to_korean(client: OpenAI, text: str) -> str:
    """Translate text to Korean using OpenAI API."""
    prompt = KOREAN_PROMPT.format(text=text)
    
    response = client.chat.completions.create(
        model=OPENAI_MODEL,
        messages=[{"role": "user", "content": prompt}],
        max_completion_tokens=2000
    )
    
    return response.choices[0].message.content.strip()


def translate_with_claude(client: Anthropic, text: str, lang_code: str) -> str:
    """Translate text to Chinese or Spanish using Claude API."""
    prompt = CLAUDE_PROMPTS[lang_code].format(text=text)
    
    message = client.messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=2048,
        messages=[{"role": "user", "content": prompt}]
    )
    
    return message.content[0].text.strip()


# ============================================================================
# CORE PIPELINE FUNCTIONS
# ============================================================================

def extract_single_file(
    anthropic_client: Anthropic,
    transcript_path: Path,
    output_dir: Path,
    prompt: str,
    speaker_profiles: dict,
    logger: logging.Logger
) -> tuple[int, str]:
    """Extract quotes from a single transcript file. Returns (quote_count, output_filename)."""
    transcript = load_file(transcript_path)
    quotes = extract_quotes_api(anthropic_client, transcript, prompt)
    quotes = enrich_quotes_with_speaker_info(quotes, speaker_profiles)
    
    output_filename = transcript_path.stem + "_quotes.json"
    output_path = output_dir / output_filename
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(quotes, f, indent=2, ensure_ascii=False)
    
    return len(quotes), output_filename


def translate_single_file(
    openai_client: OpenAI,
    anthropic_client: Anthropic,
    quotes_path: Path,
    logger: logging.Logger
) -> int:
    """Translate all quotes in a single file. Returns quote count."""
    with open(quotes_path, "r", encoding="utf-8") as f:
        quotes = json.load(f)
    
    for quote in quotes:
        text = quote["text"]
        
        # Korean (OpenAI)
        quote["text_ko"] = translate_to_korean(openai_client, text)
        time.sleep(API_DELAY_SECONDS)
        
        # Chinese (Claude)
        quote["text_zh"] = translate_with_claude(anthropic_client, text, "zh")
        time.sleep(API_DELAY_SECONDS)
        
        # Spanish (Claude)
        quote["text_es"] = translate_with_claude(anthropic_client, text, "es")
        time.sleep(API_DELAY_SECONDS)
    
    # Save translated quotes back
    with open(quotes_path, "w", encoding="utf-8") as f:
        json.dump(quotes, f, indent=2, ensure_ascii=False)
    
    return len(quotes)


def process_batch_extract(
    batch_files: list[Path],
    anthropic_client: Anthropic,
    output_dir: Path,
    prompt: str,
    speaker_profiles: dict,
    checkpoint: CheckpointManager,
    logger: logging.Logger,
    batch_num: int,
    total_batches: int,
    total_files: int,
    files_processed_before: int
) -> int:
    """Process a batch of files for extraction. Returns files processed in this batch."""
    batch_processed = 0
    
    for idx, transcript_path in enumerate(batch_files):
        filename = transcript_path.name
        global_idx = files_processed_before + idx + 1
        
        # Skip if already extracted
        if checkpoint.is_extracted(filename):
            batch_processed += 1
            continue
        
        # Progress bar
        print_progress_bar(
            global_idx, total_files,
            prefix=f"Extract B{batch_num}/{total_batches}",
            suffix=f"| {filename[:30]}..."
        )
        
        retry_count = 0
        while retry_count < MAX_RETRIES:
            try:
                quote_count, output_file = extract_single_file(
                    anthropic_client, transcript_path, output_dir,
                    prompt, speaker_profiles, logger
                )
                
                checkpoint.mark_extracted(filename, quote_count, COST_PER_FILE_EXTRACT_USD)
                batch_processed += 1
                time.sleep(API_DELAY_SECONDS)
                break
                
            except Exception as e:
                retry_count += 1
                error_msg = f"Extract {filename} (attempt {retry_count}): {str(e)}"
                logger.error(error_msg)
                
                if retry_count < MAX_RETRIES:
                    print(f"\n  -> Error, retrying in {ERROR_RETRY_WAIT_MINUTES} minutes...")
                    time.sleep(ERROR_RETRY_WAIT_MINUTES * 60)
                else:
                    checkpoint.add_error(f"FAILED after {MAX_RETRIES} attempts: {error_msg}")
                    print(f"\n  -> Failed after {MAX_RETRIES} attempts, skipping...")
    
    print()  # New line after progress bar
    return batch_processed


def process_batch_translate(
    batch_files: list[str],
    openai_client: OpenAI,
    anthropic_client: Anthropic,
    output_dir: Path,
    checkpoint: CheckpointManager,
    logger: logging.Logger,
    batch_num: int,
    total_batches: int,
    total_files: int,
    files_processed_before: int
) -> int:
    """Process a batch of files for translation. Returns files processed in this batch."""
    batch_processed = 0
    
    for idx, filename in enumerate(batch_files):
        quotes_filename = filename.replace(".txt", "_quotes.json")
        quotes_path = output_dir / quotes_filename
        global_idx = files_processed_before + idx + 1
        
        # Skip if already translated
        if checkpoint.is_translated(filename):
            batch_processed += 1
            continue
        
        # Skip if quotes file doesn't exist (extraction failed)
        if not quotes_path.exists():
            continue
        
        # Progress bar
        print_progress_bar(
            global_idx, total_files,
            prefix=f"Translate B{batch_num}/{total_batches}",
            suffix=f"| {quotes_filename[:25]}..."
        )
        
        retry_count = 0
        while retry_count < MAX_RETRIES:
            try:
                quote_count = translate_single_file(
                    openai_client, anthropic_client, quotes_path, logger
                )
                
                checkpoint.mark_translated(filename, quote_count, COST_PER_FILE_TRANSLATE_USD)
                batch_processed += 1
                break
                
            except Exception as e:
                retry_count += 1
                error_msg = f"Translate {quotes_filename} (attempt {retry_count}): {str(e)}"
                logger.error(error_msg)
                
                if retry_count < MAX_RETRIES:
                    print(f"\n  -> Error, retrying in {ERROR_RETRY_WAIT_MINUTES} minutes...")
                    time.sleep(ERROR_RETRY_WAIT_MINUTES * 60)
                else:
                    checkpoint.add_error(f"FAILED after {MAX_RETRIES} attempts: {error_msg}")
                    print(f"\n  -> Failed after {MAX_RETRIES} attempts, skipping...")
    
    print()  # New line after progress bar
    return batch_processed


def print_batch_summary(
    batch_num: int,
    files_completed: int,
    total_files: int,
    checkpoint: CheckpointManager
):
    """Print summary after each batch."""
    elapsed = checkpoint.get_elapsed_time()
    elapsed_secs = elapsed.total_seconds()
    
    # Calculate ETA
    if files_completed > 0:
        secs_per_file = elapsed_secs / files_completed
        remaining_files = total_files - files_completed
        eta_secs = remaining_files * secs_per_file
    else:
        eta_secs = 0
    
    percent = (files_completed / total_files * 100) if total_files > 0 else 0
    cost = checkpoint.data["total_cost_usd"]
    
    print("\n" + "-" * 60)
    print(f"BATCH {batch_num} COMPLETE")
    print("-" * 60)
    print(f"Progress: {files_completed}/{total_files} files ({percent:.1f}%)")
    print(f"Quotes extracted: {checkpoint.data['total_quotes_extracted']}")
    print(f"Quotes translated: {checkpoint.data['total_quotes_translated']}")
    print(f"Time elapsed: {format_time(elapsed_secs)}")
    print(f"Estimated remaining: {format_time(eta_secs)}")
    print(f"Cost so far: ${cost:.2f}")
    print("-" * 60 + "\n")


def run_export_to_excel(base_dir: Path, logger: logging.Logger):
    """Run the export_to_excel.py script functionality."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from collections import Counter
    
    output_dir = base_dir / "output"
    
    # Load all quotes
    all_quotes = []
    quote_files = sorted(output_dir.glob("*_quotes.json"))
    
    for file_path in quote_files:
        with open(file_path, "r", encoding="utf-8") as f:
            quotes = json.load(f)
            all_quotes.extend(quotes)
    
    if not all_quotes:
        logger.error("No quotes found for export")
        return
    
    # Transform to DataFrame
    rows = []
    for idx, quote in enumerate(all_quotes, start=1):
        row = {
            "id": idx,
            "speaker": quote.get("speaker", ""),
            "speaker_function": quote.get("speaker_function", ""),
            "speaker_expertise": ", ".join(quote.get("speaker_expertise", [])),
            "text": quote.get("text", ""),
            "text_ko": quote.get("text_ko", ""),
            "text_zh": quote.get("text_zh", ""),
            "text_es": quote.get("text_es", ""),
            "timestamp": quote.get("timestamp", ""),
            "context": quote.get("context", ""),
            "vocabulary_highlights": ", ".join(quote.get("vocabulary_highlights", [])),
            "topics": ", ".join(quote.get("topics", [])),
            "difficulty_level": quote.get("difficulty_level", "")
        }
        rows.append(row)
    
    columns = [
        "id", "speaker", "speaker_function", "speaker_expertise",
        "text", "text_ko", "text_zh", "text_es",
        "timestamp", "context", "vocabulary_highlights", "topics", "difficulty_level"
    ]
    
    df = pd.DataFrame(rows, columns=columns)
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [5, 6, 7, 8, 10]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-width
    max_widths = {
        "id": 6, "speaker": 25, "speaker_function": 18, "speaker_expertise": 35,
        "text": 60, "text_ko": 50, "text_zh": 50, "text_es": 50,
        "timestamp": 12, "context": 60, "vocabulary_highlights": 35,
        "topics": 30, "difficulty_level": 15
    }
    
    for col_idx, column in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        width = max_widths.get(column, 50)
        ws.column_dimensions[col_letter].width = width
    
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    
    # Save files
    excel_path = output_dir / "quotes_complete.xlsx"
    wb.save(excel_path)
    print(f"  -> Saved: {excel_path}")
    
    csv_path = output_dir / "quotes_complete.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"  -> Saved: {csv_path}")
    
    return len(all_quotes), df


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def main():
    base_dir = Path(__file__).parent
    transcripts_dir = base_dir / "transcripts"
    output_dir = base_dir / "output"
    prompt_path = base_dir / "extraction_prompt.txt"
    speaker_profiles_path = base_dir / "speaker_profiles.json"
    
    # Setup
    logger = setup_logging(base_dir)
    checkpoint = CheckpointManager(base_dir)
    output_dir.mkdir(exist_ok=True)
    
    print("\n" + "=" * 60)
    print("LENNY'S PODCAST QUOTE PIPELINE")
    print("=" * 60)
    
    # Step 1: Count transcript files
    transcript_files = sorted(transcripts_dir.glob("*.txt"))
    total_files = len(transcript_files)
    
    if total_files == 0:
        print("No .txt files found in transcripts/ folder.")
        return
    
    # Check for existing progress
    already_extracted = len(checkpoint.data["extracted_files"])
    already_translated = len(checkpoint.data["translated_files"])
    
    if already_extracted > 0 or already_translated > 0:
        print(f"\nFound existing progress:")
        print(f"  - Files extracted: {already_extracted}/{total_files}")
        print(f"  - Files translated: {already_translated}/{total_files}")
        print(f"  - Cost so far: ${checkpoint.data['total_cost_usd']:.2f}")
        
        resume = input("\nResume from checkpoint? (y/n): ").strip().lower()
        if resume != 'y':
            clear = input("Clear checkpoint and start fresh? (y/n): ").strip().lower()
            if clear == 'y':
                checkpoint.clear()
                already_extracted = 0
                already_translated = 0
            else:
                print("Aborted.")
                return
    
    # Step 2: Calculate estimates
    remaining_extract = total_files - already_extracted
    remaining_translate = total_files - already_translated
    
    estimated_cost = (remaining_extract * COST_PER_FILE_EXTRACT_USD + 
                      remaining_translate * COST_PER_FILE_TRANSLATE_USD)
    
    extract_time = remaining_extract * SECONDS_PER_FILE_EXTRACT
    translate_time = remaining_translate * AVG_QUOTES_PER_FILE * SECONDS_PER_QUOTE_TRANSLATE
    batch_wait_time = ((remaining_extract // BATCH_SIZE) + (remaining_translate // BATCH_SIZE)) * BATCH_WAIT_MINUTES * 60
    total_time_secs = extract_time + translate_time + batch_wait_time
    
    print(f"\n{'=' * 60}")
    print("PIPELINE SUMMARY")
    print(f"{'=' * 60}")
    print(f"Transcript files found: {total_files}")
    print(f"Files to extract: {remaining_extract}")
    print(f"Files to translate: {remaining_translate}")
    print(f"Batch size: {BATCH_SIZE} files")
    print(f"Total batches: {(total_files + BATCH_SIZE - 1) // BATCH_SIZE}")
    print(f"\nEstimated cost: ${estimated_cost:.2f}")
    print(f"Estimated time: {format_time(total_time_secs)}")
    print(f"{'=' * 60}")
    
    # Step 3: Ask for confirmation
    confirm = input(f"\nProceed with pipeline? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Aborted by user.")
        return
    
    # Initialize API clients
    print("\nInitializing API clients...")
    openai_client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
    anthropic_client = Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
    
    # Load resources
    print("Loading extraction prompt...")
    prompt = load_file(prompt_path)
    
    print("Loading speaker profiles...")
    with open(speaker_profiles_path, "r", encoding="utf-8") as f:
        speaker_profiles = json.load(f)
    
    # Start timer
    checkpoint.set_start_time()
    
    # Create batches
    batch_count = (total_files + BATCH_SIZE - 1) // BATCH_SIZE
    
    print(f"\nStarting pipeline with {batch_count} batches...\n")
    
    # Process each batch
    for batch_num in range(1, batch_count + 1):
        start_idx = (batch_num - 1) * BATCH_SIZE
        end_idx = min(batch_num * BATCH_SIZE, total_files)
        batch_files = transcript_files[start_idx:end_idx]
        batch_filenames = [f.name for f in batch_files]
        
        print(f"\n{'=' * 60}")
        print(f"BATCH {batch_num}/{batch_count} (files {start_idx + 1}-{end_idx})")
        print(f"{'=' * 60}\n")
        
        # Phase 1: Extract quotes
        print("Phase 1: Extracting quotes...")
        process_batch_extract(
            batch_files, anthropic_client, output_dir, prompt, speaker_profiles,
            checkpoint, logger, batch_num, batch_count, total_files, start_idx
        )
        
        # Wait between extraction and translation
        if remaining_extract > 0:
            print(f"Waiting {BATCH_WAIT_MINUTES} minutes before translation (rate limit protection)...")
            for remaining in range(BATCH_WAIT_MINUTES * 60, 0, -30):
                print(f"  {remaining // 60}m {remaining % 60}s remaining...", end="\r")
                time.sleep(30)
            print(" " * 40, end="\r")  # Clear line
        
        # Phase 2: Translate quotes
        print("Phase 2: Translating quotes...")
        process_batch_translate(
            batch_filenames, openai_client, anthropic_client, output_dir,
            checkpoint, logger, batch_num, batch_count, total_files, start_idx
        )
        
        # Print batch summary
        files_completed = len(checkpoint.data["extracted_files"])
        print_batch_summary(batch_num, files_completed, total_files, checkpoint)
        
        # Wait between batches (except last one)
        if batch_num < batch_count:
            print(f"Waiting {BATCH_WAIT_MINUTES} minutes before next batch...")
            for remaining in range(BATCH_WAIT_MINUTES * 60, 0, -30):
                print(f"  {remaining // 60}m {remaining % 60}s remaining...", end="\r")
                time.sleep(30)
            print(" " * 40, end="\r")
    
    # Step 4: Export to Excel
    print("\n" + "=" * 60)
    print("EXPORTING TO EXCEL")
    print("=" * 60)
    
    try:
        total_quotes, df = run_export_to_excel(base_dir, logger)
        print(f"\nExported {total_quotes} quotes successfully!")
    except Exception as e:
        logger.error(f"Export failed: {str(e)}")
        print(f"Export failed: {str(e)}")
    
    # Step 5: Final summary
    elapsed = checkpoint.get_elapsed_time()
    errors = checkpoint.data["errors"]
    
    print("\n" + "=" * 60)
    print("PIPELINE COMPLETE!")
    print("=" * 60)
    print(f"Total files processed: {total_files}")
    print(f"Total quotes extracted: {checkpoint.data['total_quotes_extracted']}")
    print(f"Total quotes translated: {checkpoint.data['total_quotes_translated']}")
    print(f"Total time: {format_time(elapsed.total_seconds())}")
    print(f"Total cost: ${checkpoint.data['total_cost_usd']:.2f}")
    
    if errors:
        print(f"\nErrors encountered ({len(errors)}):")
        for error in errors[-10:]:  # Show last 10 errors
            print(f"  - {error}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more (see errors.log)")
    else:
        print("\nNo errors encountered!")
    
    print("\nOutput files:")
    print(f"  - {output_dir / 'quotes_complete.xlsx'}")
    print(f"  - {output_dir / 'quotes_complete.csv'}")
    print(f"  - {len(list(output_dir.glob('*_quotes.json')))} individual JSON files")
    print("=" * 60)
    
    # Clean up checkpoint on success
    if not errors:
        cleanup = input("\nDelete checkpoint file? (y/n): ").strip().lower()
        if cleanup == 'y':
            checkpoint.checkpoint_path.unlink(missing_ok=True)
            print("Checkpoint deleted.")


if __name__ == "__main__":
    main()
